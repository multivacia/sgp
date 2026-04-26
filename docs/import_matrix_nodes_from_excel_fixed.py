from __future__ import annotations

import argparse
import json
import os
import re
import sys
import uuid
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

try:
    import psycopg  # type: ignore
    _DRIVER = "psycopg"
except Exception:  # pragma: no cover
    psycopg = None
    try:
        import psycopg2  # type: ignore
        from psycopg2.extras import Json as Psycopg2Json  # type: ignore
        _DRIVER = "psycopg2"
    except Exception:  # pragma: no cover
        psycopg2 = None
        Psycopg2Json = None
        _DRIVER = "none"

SHEET_DEFAULT = "ATIVIDADES E APONTAMENTOS"


@dataclass
class ActivityRow:
    row_number: int
    task_name: str
    sector_name: str
    activity_name: str
    responsible_name: str | None
    planned_minutes: int | None
    real_minutes: int | None
    pct_conclusao: float | None


@dataclass
class ImportContext:
    workbook_path: Path
    workbook_name: str
    sheet_name: str
    item_code: str
    item_name: str
    item_description: str | None
    dry_run: bool
    replace_existing: bool
    root_id: uuid.UUID


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    if not text or text.upper() in {"#N/A", "#N/D", "N/A", "NONE", "NULL"}:
        return None
    return text


def to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(round(float(value)))
    except Exception:
        return None


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


_slug_invalid_re = re.compile(r"[^a-z0-9]+")


def slugify(value: str) -> str:
    import unicodedata

    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    lowered = normalized.lower().strip()
    lowered = _slug_invalid_re.sub("-", lowered)
    return lowered.strip("-") or "node"


class Database:
    def __init__(self, dsn: str):
        self.dsn = dsn
        self.conn = None

    def __enter__(self):
        if _DRIVER == "psycopg":
            self.conn = psycopg.connect(self.dsn)
        elif _DRIVER == "psycopg2":
            self.conn = psycopg2.connect(self.dsn)
        else:  # pragma: no cover
            raise RuntimeError(
                "Nenhum driver PostgreSQL encontrado. Instale `psycopg[binary]` ou `psycopg2-binary`."
            )
        return self

    def __exit__(self, exc_type, exc, tb):
        if self.conn is None:
            return
        if exc_type is None:
            self.conn.commit()
        else:
            self.conn.rollback()
        self.conn.close()

    def execute(self, sql: str, params: tuple | dict | None = None):
        with self.conn.cursor() as cur:
            cur.execute(sql, params)
            return cur

    def fetchone(self, sql: str, params: tuple | dict | None = None):
        with self.conn.cursor() as cur:
            cur.execute(sql, params)
            return cur.fetchone()


class Importer:
    def __init__(self, db: Database, ctx: ImportContext):
        self.db = db
        self.ctx = ctx
        self.created_counts = {"ITEM": 0, "TASK": 0, "SECTOR": 0, "ACTIVITY": 0}
        self.collaborator_cache: dict[str, uuid.UUID | None] = {}

    def ensure_no_existing_root(self) -> None:
        row = self.db.fetchone(
            """
            SELECT id
            FROM matrix_nodes
            WHERE node_type = 'ITEM'
              AND code = %s
              AND deleted_at IS NULL
            LIMIT 1
            """,
            (self.ctx.item_code,),
        )
        if row and not self.ctx.replace_existing:
            raise RuntimeError(
                f"Já existe ITEM ativo com code={self.ctx.item_code!r}. Use --replace-existing para recarregar."
            )
        if row and self.ctx.replace_existing:
            existing_root_id = row[0]
            self.db.execute(
                """
                UPDATE matrix_nodes
                   SET deleted_at = NOW(),
                       is_active = FALSE,
                       updated_at = NOW()
                 WHERE root_id = %s
                   AND deleted_at IS NULL
                """,
                (existing_root_id,),
            )
            print(f"[info] Árvore anterior marcada como removida logicamente: root_id={existing_root_id}")

    def resolve_collaborator_id(self, full_name: str | None) -> uuid.UUID | None:
        if not full_name:
            return None
        key = full_name.strip().lower()
        if key in self.collaborator_cache:
            return self.collaborator_cache[key]

        row = self.db.fetchone(
            """
            SELECT id
            FROM collaborators
            WHERE lower(trim(full_name)) = lower(trim(%s))
              AND deleted_at IS NULL
            LIMIT 1
            """,
            (full_name,),
        )
        collaborator_id = row[0] if row else None
        self.collaborator_cache[key] = collaborator_id
        return collaborator_id

    def insert_node(
        self,
        *,
        node_id: uuid.UUID,
        parent_id: uuid.UUID | None,
        root_id: uuid.UUID,
        node_type: str,
        code: str | None,
        name: str,
        description: str | None,
        order_index: int,
        level_depth: int,
        planned_minutes: int | None = None,
        default_responsible_id: uuid.UUID | None = None,
        required: bool = True,
        metadata: dict[str, Any] | None = None,
        source_key: str | None = None,
    ) -> None:
        metadata_param: Any = metadata
        if _DRIVER == "psycopg2" and metadata is not None:
            metadata_param = Psycopg2Json(metadata)

        self.db.execute(
            """
            INSERT INTO matrix_nodes (
              id,
              parent_id,
              root_id,
              node_type,
              code,
              name,
              description,
              order_index,
              level_depth,
              is_active,
              planned_minutes,
              default_responsible_id,
              required,
              source_key,
              metadata_json,
              created_at,
              updated_at,
              deleted_at
            ) VALUES (
              %s, %s, %s, %s, %s, %s, %s,
              %s, %s, TRUE, %s, %s, %s, %s, %s,
              NOW(), NOW(), NULL
            )
            """,
            (
                str(node_id),
                str(parent_id) if parent_id else None,
                str(root_id),
                node_type,
                code,
                name,
                description,
                order_index,
                level_depth,
                planned_minutes,
                str(default_responsible_id) if default_responsible_id else None,
                required,
                source_key,
                metadata_param,
            ),
        )
        self.created_counts[node_type] += 1

    def run(self, rows: list[ActivityRow]) -> None:
        self.ensure_no_existing_root()

        root_id = self.ctx.root_id
        self.insert_node(
            node_id=root_id,
            parent_id=None,
            root_id=root_id,
            node_type="ITEM",
            code=self.ctx.item_code,
            name=self.ctx.item_name,
            description=self.ctx.item_description,
            order_index=0,
            level_depth=0,
            source_key=f"legacy-xlsx:item:{self.ctx.item_code}",
            metadata={
                "source": {
                    "workbook": self.ctx.workbook_name,
                    "sheet": self.ctx.sheet_name,
                },
                "import": {
                    "mode": "excel_dna_sgp",
                },
            },
        )

        task_nodes: OrderedDict[str, uuid.UUID] = OrderedDict()
        sector_nodes: OrderedDict[tuple[str, str], uuid.UUID] = OrderedDict()
        activity_order_per_sector: dict[uuid.UUID, int] = {}

        for activity in rows:
            if activity.task_name not in task_nodes:
                task_id = uuid.uuid4()
                task_nodes[activity.task_name] = task_id
                self.insert_node(
                    node_id=task_id,
                    parent_id=root_id,
                    root_id=root_id,
                    node_type="TASK",
                    code=f"TASK-{len(task_nodes):03d}",
                    name=activity.task_name,
                    description=None,
                    order_index=len(task_nodes) - 1,
                    level_depth=1,
                    source_key=f"legacy-xlsx:task:{slugify(activity.task_name)}",
                    metadata={
                        "source": {
                            "workbook": self.ctx.workbook_name,
                            "sheet": self.ctx.sheet_name,
                            "first_excel_row": activity.row_number,
                        }
                    },
                )

            task_id = task_nodes[activity.task_name]
            sector_key = (activity.task_name, activity.sector_name)
            if sector_key not in sector_nodes:
                sector_id = uuid.uuid4()
                sector_nodes[sector_key] = sector_id
                sector_order_index = sum(1 for t, _ in sector_nodes.keys() if t == activity.task_name) - 1
                self.insert_node(
                    node_id=sector_id,
                    parent_id=task_id,
                    root_id=root_id,
                    node_type="SECTOR",
                    code=f"SECTOR-{slugify(activity.sector_name)}",
                    name=activity.sector_name,
                    description=None,
                    order_index=sector_order_index,
                    level_depth=2,
                    source_key=(
                        f"legacy-xlsx:sector:{slugify(activity.task_name)}:{slugify(activity.sector_name)}"
                    ),
                    metadata={
                        "source": {
                            "workbook": self.ctx.workbook_name,
                            "sheet": self.ctx.sheet_name,
                            "first_excel_row": activity.row_number,
                        }
                    },
                )
                activity_order_per_sector[sector_id] = 0

            sector_id = sector_nodes[sector_key]
            activity_order_index = activity_order_per_sector[sector_id]
            activity_order_per_sector[sector_id] += 1
            collaborator_id = self.resolve_collaborator_id(activity.responsible_name)
            if activity.responsible_name and collaborator_id is None:
                print(
                    f"[warn] Colaborador não encontrado em collaborators: {activity.responsible_name!r} "
                    f"(linha Excel {activity.row_number})"
                )

            metadata = {
                "source": {
                    "workbook": self.ctx.workbook_name,
                    "sheet": self.ctx.sheet_name,
                    "excel_row": activity.row_number,
                },
                "legacy": {
                    "responsible_name": activity.responsible_name,
                    "real_minutes": activity.real_minutes,
                    "pct_conclusao": activity.pct_conclusao,
                },
            }
            self.insert_node(
                node_id=uuid.uuid4(),
                parent_id=sector_id,
                root_id=root_id,
                node_type="ACTIVITY",
                code=None,
                name=activity.activity_name,
                description=None,
                order_index=activity_order_index,
                level_depth=3,
                planned_minutes=activity.planned_minutes,
                default_responsible_id=collaborator_id,
                required=True,
                source_key=f"legacy-xlsx:activity:{self.ctx.item_code}:{activity.row_number}",
                metadata=metadata,
            )

    def print_summary(self) -> None:
        print("[ok] Importação concluída.")
        print(json.dumps(self.created_counts, ensure_ascii=False, indent=2))


def build_dsn_from_env() -> str:
    database_url = os.getenv("DATABASE_URL")
    if database_url:
        return database_url

    host = os.getenv("PGHOST")
    port = os.getenv("PGPORT")
    dbname = os.getenv("PGDATABASE")
    user = os.getenv("PGUSER")
    password = os.getenv("PGPASSWORD")

    missing = [
        name
        for name, value in {
            "PGHOST": host,
            "PGPORT": port,
            "PGDATABASE": dbname,
            "PGUSER": user,
            "PGPASSWORD": password,
        }.items()
        if not value
    ]
    if missing:
        raise RuntimeError(
            "Configuração PostgreSQL ausente. Informe DATABASE_URL ou todas as variáveis PG*. "
            f"Faltando: {', '.join(missing)}"
        )

    def escape_conninfo(value: str) -> str:
        return value.replace("\", "\\").replace("'", "\'")

    return (
        f"host='{escape_conninfo(host)}' "
        f"port='{escape_conninfo(port)}' "
        f"dbname='{escape_conninfo(dbname)}' "
        f"user='{escape_conninfo(user)}' "
        f"password='{escape_conninfo(password)}'"
    )


def read_activity_rows(workbook_path: Path, sheet_name: str) -> list[ActivityRow]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Aba {sheet_name!r} não encontrada. Abas disponíveis: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows: list[ActivityRow] = []
    for row_number in range(2, ws.max_row + 1):
        task_name = normalize_text(ws.cell(row_number, 2).value)
        sector_name = normalize_text(ws.cell(row_number, 3).value)
        activity_name = normalize_text(ws.cell(row_number, 4).value)
        responsible_name = normalize_text(ws.cell(row_number, 5).value)
        planned_minutes = to_int(ws.cell(row_number, 6).value)
        real_minutes = to_int(ws.cell(row_number, 7).value)
        pct_conclusao = to_float(ws.cell(row_number, 8).value)

        if not task_name or not sector_name or not activity_name:
            continue

        rows.append(
            ActivityRow(
                row_number=row_number,
                task_name=task_name,
                sector_name=sector_name,
                activity_name=activity_name,
                responsible_name=responsible_name,
                planned_minutes=planned_minutes,
                real_minutes=real_minutes,
                pct_conclusao=pct_conclusao,
            )
        )

    if not rows:
        raise RuntimeError("Nenhuma atividade válida encontrada no Excel.")
    return rows


def infer_item_name_from_filename(workbook_path: Path) -> str:
    stem = workbook_path.stem
    return re.sub(r"\s+", " ", stem.replace("_", " ")).strip()


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa Excel legado do SGP para matrix_nodes (ITEM -> TASK -> SECTOR -> ACTIVITY)."
    )
    parser.add_argument("xlsx", type=Path, help="Caminho do arquivo .xlsx")
    parser.add_argument("--sheet", default=SHEET_DEFAULT, help=f"Aba a ler (default: {SHEET_DEFAULT})")
    parser.add_argument("--item-code", required=True, help="Código único do ITEM raiz, ex.: ITEM-CARPETE")
    parser.add_argument("--item-name", help="Nome do ITEM raiz. Se omitido, tenta inferir do nome do arquivo.")
    parser.add_argument("--item-description", help="Descrição opcional para o ITEM raiz.")
    parser.add_argument(
        "--replace-existing",
        action="store_true",
        help="Se já existir ITEM ativo com o mesmo code, marca a árvore anterior como removida logicamente e recarrega.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Lê e valida o Excel, imprime resumo e não grava no banco.",
    )
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    workbook_path: Path = args.xlsx
    if not workbook_path.exists():
        raise RuntimeError(f"Arquivo não encontrado: {workbook_path}")

    rows = read_activity_rows(workbook_path, args.sheet)
    task_count = len(OrderedDict((row.task_name, True) for row in rows))
    sector_count = len(OrderedDict(((row.task_name, row.sector_name), True) for row in rows))
    activity_count = len(rows)

    item_name = args.item_name or infer_item_name_from_filename(workbook_path)
    root_id = uuid.uuid4()
    ctx = ImportContext(
        workbook_path=workbook_path,
        workbook_name=workbook_path.name,
        sheet_name=args.sheet,
        item_code=args.item_code,
        item_name=item_name,
        item_description=args.item_description,
        dry_run=args.dry_run,
        replace_existing=args.replace_existing,
        root_id=root_id,
    )

    print("[info] Prévia da carga")
    print(
        json.dumps(
            {
                "workbook": workbook_path.name,
                "sheet": args.sheet,
                "item_code": args.item_code,
                "item_name": item_name,
                "tasks": task_count,
                "task_sector_pairs": sector_count,
                "activities": activity_count,
                "dry_run": args.dry_run,
            },
            ensure_ascii=False,
            indent=2,
        )
    )

    if args.dry_run:
        return 0

    dsn = build_dsn_from_env()
    with Database(dsn) as db:
        importer = Importer(db, ctx)
        importer.run(rows)
        importer.print_summary()
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except Exception as exc:
        print(f"[erro] {exc}", file=sys.stderr)
        raise
